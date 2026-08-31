import sqlite3
import streamlit as st
import pandas as pd
import plotly.graph_objects as go
import openpyxl
from pathlib import Path
from datetime import datetime, timedelta

FORECAST_HORIZON_DAYS = 10  # business days
FORECAST_CI = 0.80  # 80% prediction interval

# ── JSA Brand Colors (aligned to the Admin Portal shell's shared palette) ─────
JPSI_DARK = "#32373c"
JPSI_BLUE = "#0693e3"
BG        = "#f6f8f7"
CARD_BG   = "#ffffff"
SURFACE2  = "#eef3f0"
BORDER    = "#d7e2dc"
MUTED     = "#5f7267"
TEXT      = "#32373c"
POS       = "#16a34a"
NEG       = "#dc2626"
NEU       = "#5f7267"

JSA_LOGO = "https://www.jpsi.com/wp-content/themes/gate39media/img/logo-full.png"
DATA_PATH = Path(__file__).parent / "data" / "Feeder Cattle Info Ross.xlsx"
MARS_DB_PATH = Path(__file__).parent / "data" / "mars_history.db"

VALID_STATES = {"CO", "IA", "KS", "MO", "MT", "NE", "NM", "ND", "OK", "SD", "TX", "WY"}

# st.set_page_config removed — the JSA Admin Portal shell (Home.py) makes the
# single set_page_config call allowed per multi-page run.

st.markdown(f"""
<style>
  html, body, [data-testid="stAppViewContainer"] {{
    background-color:{BG}; color:{TEXT};
  }}
  [data-testid="stSidebar"] {{
    background-color:{SURFACE2}; border-right:1px solid {BORDER};
  }}
  .tile {{
    background:{CARD_BG}; border:1px solid {BORDER};
    border-top:3px solid {JPSI_BLUE}; border-radius:10px;
    padding:16px 20px; text-align:center; height:100%;
    box-shadow: 0 1px 2px rgba(0,0,0,0.04);
  }}
  .tile-label {{
    color:{MUTED}; font-size:0.68rem; text-transform:uppercase;
    letter-spacing:0.09em; margin-bottom:6px;
  }}
  .tile-value {{
    color:{TEXT}; font-size:1.7rem; font-weight:700; line-height:1.1;
  }}
  .tile-delta-pos {{ color:{POS}; font-size:0.82rem; font-weight:600; margin-top:4px; }}
  .tile-delta-neg {{ color:{NEG}; font-size:0.82rem; font-weight:600; margin-top:4px; }}
  .tile-delta-neu {{ color:{NEU}; font-size:0.82rem; font-weight:600; margin-top:4px; }}
  .sec-header {{
    color:{MUTED}; font-size:0.72rem; text-transform:uppercase;
    letter-spacing:0.1em; padding:8px 0 4px; border-bottom:1px solid {BORDER};
    margin-bottom:10px;
  }}
  hr {{ border-color:{BORDER}; }}
  #MainMenu, footer {{ visibility:hidden; }}
  .stDeployButton {{ display:none; }}
  div[class*="st-key-wm-"] {{ position:relative; }}
  div[class*="st-key-wm-"]::after {{
    content:"";
    position:absolute; inset:0; margin:auto;
    width:50%; height:110px; max-width:320px;
    background-image:url('{JSA_LOGO}');
    background-size:contain; background-repeat:no-repeat; background-position:center;
    opacity:0.05; pointer-events:none; z-index:3;
  }}
</style>
""", unsafe_allow_html=True)


# ── Helpers ───────────────────────────────────────────────────────────────────

def delta_html(val, suffix=""):
    if val is None or pd.isna(val):
        return '<div class="tile-delta-neu">—</div>'
    sign = "▲" if val > 0 else ("▼" if val < 0 else "")
    color = "pos" if val > 0 else ("neg" if val < 0 else "neu")
    return f'<div class="tile-delta-{color}">{sign} ${abs(val):.2f}{suffix}</div>'


def tile(label, value, delta=""):
    return (f'<div class="tile">'
            f'<div class="tile-label">{label}</div>'
            f'<div class="tile-value">{value}</div>'
            f'{delta}</div>')


def fmt_price(v):
    if v is None or pd.isna(v):
        return "—"
    return f"-${abs(v):.2f}" if v < 0 else f"${v:.2f}"


def value_on_or_before(df, target_date):
    """Latest fci_value at or before target_date."""
    sub = df[df["date"] <= target_date]
    return sub.iloc[-1]["fci_value"] if not sub.empty else None


def add_watermark(fig, size=0.32, opacity=0.06):
    fig.add_layout_image(dict(
        source=JSA_LOGO,
        xref="paper", yref="paper", x=0.5, y=0.5,
        xanchor="center", yanchor="middle",
        sizex=size, sizey=size,
        opacity=opacity, layer="below",
    ))
    return fig


@st.cache_data(ttl=3600, show_spinner=False)
def compute_forecast(fci_values, last_date, horizon=FORECAST_HORIZON_DAYS, ci=FORECAST_CI):
    """
    Naive (random-walk) trend projection, flat at the current value, with a
    band built from the historical distribution of actual h-day-ahead price
    changes -- not a fitted statistical model. Backtested against Holt's
    exponential smoothing (66 rolling-origin trials, 2023-10 to 2026-08):
    simple "no change" persistence beat Holt at every forecast horizon
    (overall MAE $3.71 vs $4.07/cwt) -- this series behaves close to a
    random walk, where trend-extrapolation added no real edge, so the
    simpler and more accurate approach is used here instead. This is
    descriptive of typical historical variability, not a trading signal or
    market forecast. Returns None if there isn't enough history.
    """
    y = pd.Series(fci_values).astype(float).reset_index(drop=True)
    if len(y) < 60:
        return None
    current = y.iloc[-1]
    lo_q, hi_q = (1 - ci) / 2, 1 - (1 - ci) / 2
    future_dates = pd.bdate_range(start=pd.Timestamp(last_date) + pd.Timedelta(days=1), periods=horizon)
    rows = []
    for h, d in enumerate(future_dates, start=1):
        changes = (y - y.shift(h)).dropna()
        if len(changes) < 20:
            lower = upper = current
        else:
            lower = current + changes.quantile(lo_q)
            upper = current + changes.quantile(hi_q)
        rows.append({"date": d, "forecast": current, "lower": lower, "upper": upper})
    return pd.DataFrame(rows)


# ── Data Loading ──────────────────────────────────────────────────────────────

def _load_workbook():
    """
    Source: 'Sale Location Data 24-25' sheet of the JSA-compiled workbook.
    Each date has one row per reporting sale location (Daily $ price) plus a
    synthetic 'FCI' row holding that date's published CME Feeder Cattle Index
    value. Basis is recomputed here (location price - that day's FCI) so it
    stays consistent even where the source sheet's own Basis column is stale.
    Covers 2024-01-01 through 2026-01-23.
    """
    raw = pd.read_excel(DATA_PATH, sheet_name="Sale Location Data 24-25", usecols="A:D", header=0)
    raw.columns = ["date", "location", "state", "price"]
    raw = raw.dropna(subset=["date", "location", "price"])
    raw["date"] = pd.to_datetime(raw["date"])
    raw["location"] = raw["location"].astype(str).str.strip().str.upper()
    raw["state"] = raw["state"].astype(str).str.strip().str.upper()

    raw_fci = raw[raw["location"] == "FCI"][["date", "price"]].sort_values("date")
    # The sheet has exactly one duplicate-dated FCI row (2024-07-23: 258.39 vs
    # a clearly erroneous 326.18, a lone data-entry duplicate, not a pattern)
    # -- picking blindly via keep="first"/"last" is a coin flip on which one
    # survives, so instead keep whichever candidate is closest to the median
    # of the surrounding +/-5 days' (non-duplicate) values.
    dupe_dates = raw_fci[raw_fci.duplicated(subset="date", keep=False)]["date"].unique()
    if len(dupe_dates):
        single = raw_fci[~raw_fci["date"].isin(dupe_dates)].set_index("date")["price"]
        keep_rows = []
        for d in dupe_dates:
            candidates = raw_fci[raw_fci["date"] == d]
            window = single[(single.index >= d - pd.Timedelta(days=5)) & (single.index <= d + pd.Timedelta(days=5))]
            ref = window.median() if not window.empty else candidates["price"].median()
            keep_rows.append((candidates["price"] - ref).abs().idxmin())
        raw_fci = pd.concat([raw_fci[~raw_fci["date"].isin(dupe_dates)], raw_fci.loc[keep_rows]])

    fci = (
        raw_fci.rename(columns={"price": "fci_value"})
        .sort_values("date")
        .reset_index(drop=True)
    )
    fci["source"] = "workbook"
    # same-day (non-rolling) snapshot isn't in this sheet either
    fci["same_day_price"] = pd.NA
    fci["same_day_head"] = pd.NA
    fci["same_day_avg_weight"] = pd.NA

    loc = raw[raw["location"] != "FCI"].copy()
    loc = loc.merge(fci[["date", "fci_value"]], on="date", how="left")
    loc["basis"] = loc["price"] - loc["fci_value"]
    loc = loc.dropna(subset=["fci_value"])
    loc["source"] = "workbook"
    # head count / avg weight aren't in this sheet — the location table shows "—" for these dates
    loc["head"] = pd.NA
    loc["avg_weight"] = pd.NA

    return fci, loc


def _load_mars_reconstruction():
    """
    Continues the timeline past the workbook's last date (2026-01-23) using
    USDA AMS MARS sale-barn data, reconstructed with the same weighted-average
    methodology as the workbook's own 'FCI Estimation' sheet:

        FCI(date) = sum(head*weight*price) / sum(head*weight)

    across a ~60-location roster in the CME 12-state region (see
    update_index.py / data/mars_roster.json). This is JSA's own reconstruction,
    not CME's official feed, and has been spot-checked against CME's published
    values within roughly $2-9/cwt on any given day (missing Direct/Video/
    Internet trade volume, which this sale-barn-only roster doesn't capture).
    Run `python update_index.py` to refresh.
    """
    fci_cols = ["date", "fci_value", "source", "same_day_price", "same_day_head", "same_day_avg_weight"]
    loc_cols = ["date", "location", "state", "price", "head", "avg_weight", "fci_value", "basis", "source"]
    if not MARS_DB_PATH.exists():
        return pd.DataFrame(columns=fci_cols), pd.DataFrame(columns=loc_cols)

    conn = sqlite3.connect(MARS_DB_PATH)
    fci_raw = pd.read_sql(
        "SELECT report_date AS date, fci_value, same_day_price, same_day_head, same_day_avg_weight "
        "FROM fci_daily", conn,
    )
    sales = pd.read_sql(
        "SELECT report_date AS date, location, state, weight_low, head_count, avg_weight, avg_price "
        "FROM mars_sales", conn,
    )
    conn.close()

    fci = fci_raw.copy()
    fci["date"] = pd.to_datetime(fci["date"])
    fci["source"] = "usda_mars"
    fci = fci.sort_values("date").reset_index(drop=True)

    if sales.empty:
        loc = pd.DataFrame(columns=loc_cols)
        return fci, loc

    sales["date"] = pd.to_datetime(sales["date"])
    sales["w"] = sales["head_count"] * sales["avg_weight"]
    sales["wp"] = sales["w"] * sales["avg_price"]
    daily_loc = (
        sales.groupby(["date", "location", "state"])
        .agg(w=("w", "sum"), wp=("wp", "sum"), head=("head_count", "sum"))
        .reset_index()
    )
    daily_loc["price"] = daily_loc["wp"] / daily_loc["w"]
    # weighted-average weight across whatever brackets/grades a location reported that day
    daily_loc["avg_weight"] = daily_loc["w"] / daily_loc["head"]
    loc = daily_loc.merge(fci[["date", "fci_value"]], on="date", how="left")
    loc["basis"] = loc["price"] - loc["fci_value"]
    loc["source"] = "usda_mars"
    loc = loc.dropna(subset=["fci_value"])[loc_cols]

    return fci, loc


WB_PRECURSOR_SHEET = "CME Feeder Cattle Index Values"
BRACKET_SPECS = [
    (700, "1"), (750, "1"), (800, "1"), (850, "1"),
    (700, "1-2"), (750, "1-2"), (800, "1-2"), (850, "1-2"),
]


def _parse_bracket_cell(cell):
    """
    Cells hold 'head weight price' as three whitespace-separated numbers, but
    weight and price are sometimes concatenated with no separator when weight
    has decimals (e.g. '172  812.90257.25') -- always exactly two 6-char
    'DDD.DD' halves in that case (weight ~700-899 lbs, price ~$150-450/cwt,
    both always 3 integer digits + 2 decimals in this sheet). Cells with no
    sale in that bracket are '0   0   0.00' or a literal '//////' placeholder.
    """
    if cell is None:
        return None
    parts = str(cell).strip().split()
    try:
        if len(parts) == 3:
            head, weight, price = int(float(parts[0])), float(parts[1]), float(parts[2])
        elif len(parts) == 2:
            head = int(float(parts[0]))
            if head == 0:
                return None
            merged = parts[1]
            if len(merged) != 12:
                return None
            weight, price = float(merged[:6]), float(merged[6:])
        else:
            return None
    except ValueError:
        return None
    if head <= 0 or weight <= 0 or price <= 0:
        return None
    return head, weight, price


@st.cache_data(ttl=3600, show_spinner=False)
def _load_workbook_precursor(before_date):
    """
    'CME Feeder Cattle Index Values' is Ross's raw per-location, per-weight-
    bracket sale data (2023-01-24 - 2026-01-23) -- the actual source data
    behind the workbook's published FCI column, not a separate estimate.
    Recomputing CME's 7-day rolling weighted-average methodology from these
    raw rows reproduces the published index almost exactly (median abs error
    ~$0.003/cwt across the full 2024-2026 overlap, spot-checked 2026-08-26)
    -- far more accurate than the USDA MARS reconstruction for the same kind
    of gap, since this is the real underlying data rather than an
    approximation from a different source. Used only for dates before
    `before_date` (the main workbook sheet's own start), extending the
    ground-truth-quality range back to 2023-01-24.
    """
    fci_cols = ["date", "fci_value", "source", "same_day_price", "same_day_head", "same_day_avg_weight"]
    loc_cols = ["date", "location", "state", "price", "head", "avg_weight", "fci_value", "basis", "source"]
    try:
        wb = openpyxl.load_workbook(DATA_PATH, read_only=True, data_only=True)
        ws = wb[WB_PRECURSOR_SHEET]
    except Exception:
        return pd.DataFrame(columns=fci_cols), pd.DataFrame(columns=loc_cols)

    by_day = {}
    by_day_loc = {}
    for r in ws.iter_rows(values_only=True):
        if not isinstance(r[0], datetime):
            continue
        d, loc, state = r[0].date(), r[1], r[2]
        if not loc or not state:
            continue
        for bi in range(8):
            parsed = _parse_bracket_cell(r[3 + bi])
            if parsed is None:
                continue
            head, weight, price = parsed
            w = head * weight
            wp = w * price
            by_day.setdefault(d, []).append((w, wp, head))
            by_day_loc.setdefault((d, loc, state), []).append((w, wp, head))
    wb.close()

    if not by_day:
        return pd.DataFrame(columns=fci_cols), pd.DataFrame(columns=loc_cols)

    # Drop leading report dates isolated by a large gap from the next one --
    # a single stray early report (e.g. one location, months before dense
    # coverage resumes) isn't a real sample of the 12-state index, and
    # leaving it in would make the chart draw a straight line across the gap
    # to the next real point, fabricating a multi-month "trend" that never
    # happened. Found via real data: one 2023-01-24 report, then a 251-day
    # gap before continuous coverage starts 2023-10-02.
    STALE_GAP_DAYS = 30
    report_dates = sorted(by_day)
    while len(report_dates) >= 2 and (report_dates[1] - report_dates[0]).days > STALE_GAP_DAYS:
        stale = report_dates.pop(0)
        del by_day[stale]
    if not by_day:
        return pd.DataFrame(columns=fci_cols), pd.DataFrame(columns=loc_cols)

    before = pd.Timestamp(before_date).date()
    first_date, last_date = min(by_day), max(by_day)
    fci_rows = []
    d = first_date
    while d <= last_date:
        if d >= before:
            d += timedelta(days=1)
            continue
        window = [d - timedelta(days=i) for i in range(7)]
        num = den = 0.0
        for wd in window:
            for w, wp, head in by_day.get(wd, []):
                den += w
                num += wp
        if den <= 0:
            d += timedelta(days=1)
            continue
        sd = by_day.get(d, [])
        sd_den = sum(w for w, wp, head in sd)
        sd_num = sum(wp for w, wp, head in sd)
        sd_head = sum(head for w, wp, head in sd)
        fci_rows.append({
            "date": pd.Timestamp(d), "fci_value": num / den, "source": "workbook_precursor",
            "same_day_price": (sd_num / sd_den) if sd_den > 0 else pd.NA,
            "same_day_head": sd_head if sd_head > 0 else pd.NA,
            "same_day_avg_weight": (sd_den / sd_head) if sd_head > 0 else pd.NA,
        })
        d += timedelta(days=1)
    fci = pd.DataFrame(fci_rows, columns=fci_cols)
    if fci.empty:
        return fci, pd.DataFrame(columns=loc_cols)
    fci_by_date = dict(zip(fci["date"], fci["fci_value"]))

    loc_rows = []
    for (d, loc, state), rows in by_day_loc.items():
        if d >= before:
            continue
        ts = pd.Timestamp(d)
        if ts not in fci_by_date:
            continue
        den = sum(w for w, wp, head in rows)
        num = sum(wp for w, wp, head in rows)
        head_total = sum(head for w, wp, head in rows)
        if den <= 0 or head_total <= 0:
            continue
        price = num / den
        loc_rows.append({
            "date": ts, "location": str(loc).strip().upper(), "state": str(state).strip().upper(),
            "price": price, "head": head_total, "avg_weight": den / head_total,
            "fci_value": fci_by_date[ts], "basis": price - fci_by_date[ts], "source": "workbook_precursor",
        })
    loc = pd.DataFrame(loc_rows, columns=loc_cols)
    return fci, loc


@st.cache_data(ttl=3600, show_spinner=False)
def load_data():
    """
    Priority order, earliest ground-truth-quality data wins for each date:
      1. wb_fci (2024-01-01 - 2026-01-23): the workbook's published CME
         values -- ground truth, never overridden.
      2. precursor (2023-01-24 - 2023-12-31): recomputed from Ross's raw
         per-location sale data using CME's own methodology -- validated to
         within about $0.20/cwt (median $0.003) of the published column
         where the two overlap, so treated as ground-truth-equivalent.
      3. mars_before (before 2023-01-24) / mars_after (2026-01-24 onward):
         JSA's own USDA MARS reconstruction, same weighted-average method,
         looser accuracy (~$0.50-$9/cwt spot-checked) since it's a genuine
         approximation rather than the real underlying sale data.
    """
    wb_fci, wb_loc = _load_workbook()
    wb_start, wb_end = wb_fci["date"].min(), wb_fci["date"].max()

    precursor_fci, precursor_loc = _load_workbook_precursor(wb_start)
    precursor_start = precursor_fci["date"].min() if not precursor_fci.empty else wb_start

    mars_fci, mars_loc = _load_mars_reconstruction()
    mars_before_fci = mars_fci[mars_fci["date"] < precursor_start]
    mars_before_loc = mars_loc[mars_loc["date"] < precursor_start]
    mars_after_fci = mars_fci[mars_fci["date"] > wb_end]
    mars_after_loc = mars_loc[mars_loc["date"] > wb_end]

    fci = pd.concat(
        [mars_before_fci, precursor_fci, wb_fci, mars_after_fci], ignore_index=True
    ).sort_values("date").reset_index(drop=True)
    loc = pd.concat(
        [mars_before_loc, precursor_loc, wb_loc, mars_after_loc], ignore_index=True
    ).sort_values("date").reset_index(drop=True)

    return fci, loc


with st.spinner("Loading feeder cattle sale data…"):
    try:
        fci_df, loc_df = load_data()
        load_ok = True
        err_msg = ""
    except Exception as e:
        load_ok = False
        err_msg = str(e)
        fci_df, loc_df = pd.DataFrame(), pd.DataFrame()

if not load_ok:
    st.error("Could not load the source workbook.")
    with st.expander("Technical details"):
        st.code(err_msg)
    st.stop()

if fci_df.empty:
    st.warning("No FCI values found in the source data.")
    st.stop()

last_date = fci_df["date"].max()
first_date = fci_df["date"].min()


# ── Sidebar ───────────────────────────────────────────────────────────────────

with st.sidebar:
    st.image(JSA_LOGO, use_container_width=True)
    st.markdown("<hr>", unsafe_allow_html=True)

    st.markdown('<div class="sec-header">Detail Date</div>', unsafe_allow_html=True)
    detail_date = st.date_input(
        "Location detail for",
        value=last_date.date(),
        min_value=first_date.date(),
        max_value=last_date.date(),
        label_visibility="collapsed",
    )
    detail_date = pd.Timestamp(detail_date)

    st.markdown("<hr>", unsafe_allow_html=True)
    st.markdown('<div class="sec-header">State Filter</div>', unsafe_allow_html=True)
    all_states = sorted(loc_df["state"].unique().tolist())
    default_states = [s for s in all_states if s in VALID_STATES] or all_states
    state_filter = st.multiselect("States", all_states, default=default_states, label_visibility="collapsed")

    st.markdown("<hr>", unsafe_allow_html=True)
    if st.button("↺  Refresh Data", use_container_width=True):
        st.cache_data.clear()
        st.rerun()

    _SOURCE_LABELS = {
        "workbook": "JSA-compiled workbook, CME's published index value.",
        "workbook_precursor": "JSA reconstruction from Ross's raw per-location sale data (the same source behind the published column) — validated within about $0.20/cwt of published values where they overlap.",
        "usda_mars": "JSA reconstruction from USDA MARS sale-barn data (~60 locations), plus Direct/Video trade PDFs — spot-checked within roughly $0.50–$9/cwt of published values, not an official CME feed.",
    }
    _segments = sorted(
        (grp["date"].min(), grp["date"].max(), src) for src, grp in fci_df.groupby("source")
    )
    _seg_html = "".join(
        f'<b>{start.strftime("%b %d, %Y")} – {end.strftime("%b %d, %Y")}:</b><br>'
        f'{_SOURCE_LABELS.get(src, src)}<br><br>'
        for start, end, src in _segments
    )

    st.markdown("<hr>", unsafe_allow_html=True)
    st.markdown(
        f'<div style="color:{MUTED};font-size:0.72rem;line-height:1.6;">'
        f'Sample: 12-state feeder steer region<br>'
        f'(CO, IA, KS, MO, MT, NE, NM, ND, OK, SD, TX, WY)<br><br>'
        f'Grade/weight: #1 &amp; #1-2 Steers, Medium &amp; Large,<br>'
        f'700–899 lbs, FOB 3% standing shrink<br><br>'
        f'Coverage: {first_date.strftime("%b %d, %Y")} – {last_date.strftime("%b %d, %Y")}<br><br>'
        + _seg_html
        + f'All segments are drawn as one solid line — not<br>'
        f'visually distinguished, see this panel for which<br>'
        f'dates are which. Run <code>update_index.py</code><br>'
        f'to refresh.'
        f'</div>',
        unsafe_allow_html=True,
    )

loc_filtered = loc_df[loc_df["state"].isin(state_filter)] if state_filter else loc_df
# Guards against a stale cached load_data() result from before these columns
# existed surviving a Streamlit Cloud soft-redeploy (in-memory cache, ttl=1hr).
for _col in ("head", "avg_weight"):
    if _col not in loc_filtered.columns:
        loc_filtered[_col] = pd.NA


# ── Header ────────────────────────────────────────────────────────────────────

c1, c2 = st.columns([7, 3])
with c1:
    st.markdown(
        f"<h1 style='color:{TEXT};margin:0;padding:0;font-size:1.9rem;'>"
        "JSA - CME Feeder Cattle Index</h1>"
        f"<div style='color:{MUTED};font-size:0.8rem;margin-top:2px;'>"
        "12-State Feeder Steer Sample · #1 &amp; #1-2 Medium &amp; Large, 700–899 lbs</div>",
        unsafe_allow_html=True,
    )
with c2:
    st.markdown(
        f"<div style='text-align:right;color:{MUTED};font-size:0.75rem;padding-top:6px;'>"
        f"Most recent index date<br>"
        f"<span style='color:{JPSI_BLUE};font-size:1rem;font-weight:700;'>"
        f"{last_date.strftime('%b %d, %Y')}</span></div>",
        unsafe_allow_html=True,
    )

st.markdown("<hr style='margin:10px 0 18px;'>", unsafe_allow_html=True)


# ── KPI Tiles ─────────────────────────────────────────────────────────────────

current = fci_df.iloc[-1]["fci_value"]
prev_point = fci_df.iloc[-2]["fci_value"] if len(fci_df) > 1 else None
day_chg = current - prev_point if prev_point is not None else None

# "Current Index" is only accurate when the latest date is CME's actual
# published value (source == "workbook"). Every other date is JSA's own
# reconstruction -- label it as an estimate so it's never mistaken for the
# real published figure, which is what actually confused the user here.
current_label = (
    "Current Index" if fci_df.iloc[-1]["source"] == "workbook"
    else f"FCI Estimate {last_date.month}/{last_date.day}/{last_date.strftime('%y')}"
)

week_ago = value_on_or_before(fci_df.iloc[:-1], last_date - timedelta(days=7))
week_chg = current - week_ago if week_ago is not None else None

month_ago = value_on_or_before(fci_df.iloc[:-1], last_date - timedelta(days=30))
month_chg = current - month_ago if month_ago is not None else None

year_ago = value_on_or_before(fci_df.iloc[:-1], last_date - timedelta(days=365))
year_chg = current - year_ago if year_ago is not None else None

cols = st.columns(4)
with cols[0]:
    st.markdown(tile(current_label, fmt_price(current)), unsafe_allow_html=True)
with cols[1]:
    st.markdown(tile("Day Change", fmt_price(day_chg), delta_html(day_chg)), unsafe_allow_html=True)
with cols[2]:
    st.markdown(tile("Week Change", fmt_price(week_chg), delta_html(week_chg)), unsafe_allow_html=True)
with cols[3]:
    st.markdown(tile("Month Change", fmt_price(month_chg), delta_html(month_chg)), unsafe_allow_html=True)

# ── Daily (same-day, non-rolling) snapshot ─────────────────────────────────────
# Mirrors the "Daily: $X on Y head and Z lbs average" line under CME subscriber
# reports — the single date's own weighted average, distinct from the 7-day
# rolling Current Index above it.
last_row = fci_df.iloc[-1]
sd_price = last_row.get("same_day_price")
sd_head = last_row.get("same_day_head")
sd_weight = last_row.get("same_day_avg_weight")
if pd.notna(sd_price) and pd.notna(sd_head):
    weight_part = f" and <b style='color:{TEXT}'>{sd_weight:.0f} lbs</b> average" if pd.notna(sd_weight) else ""
    st.markdown(
        f"<div style='color:{MUTED};font-size:0.82rem;margin-top:10px;'>"
        f"Daily: <b style='color:{TEXT}'>${sd_price:.2f}</b> on "
        f"<b style='color:{TEXT}'>{int(sd_head):,}</b> head{weight_part}"
        f"</div>",
        unsafe_allow_html=True,
    )


# ── FCI Trend Chart ───────────────────────────────────────────────────────────

st.markdown("<div style='height:18px'></div>", unsafe_allow_html=True)
st.markdown('<div class="sec-header">Index Trend</div>', unsafe_allow_html=True)

AXIS = dict(
    gridcolor=BORDER, linecolor=BORDER, showgrid=True,
    tickfont=dict(color=MUTED, size=11),
    title_font=dict(color=MUTED, size=11),
    zeroline=False,
)

hover_source = fci_df["source"].map({
    "workbook": "Published",
    "workbook_precursor": "JSA reconstruction (Ross data)",
    "usda_mars": "JSA reconstruction (USDA MARS)",
})

forecast_df = compute_forecast(fci_df["fci_value"], last_date)

fig = go.Figure()
fig.add_trace(go.Scatter(
    x=fci_df["date"], y=fci_df["fci_value"],
    customdata=hover_source,
    name="CME Feeder Cattle Index", mode="lines",
    line=dict(color=JPSI_BLUE, width=2),
    hovertemplate="<b>%{customdata}</b>: $%{y:.2f}<extra></extra>",
))
if forecast_df is not None:
    connector = pd.concat([
        pd.DataFrame({"date": [last_date], "forecast": [current], "lower": [current], "upper": [current]}),
        forecast_df,
    ], ignore_index=True)
    fig.add_trace(go.Scatter(
        x=pd.concat([connector["date"], connector["date"][::-1]]),
        y=pd.concat([connector["upper"], connector["lower"][::-1]]),
        fill="toself", fillcolor="rgba(230,126,34,0.15)",
        line=dict(color="rgba(0,0,0,0)"),
        hoverinfo="skip", showlegend=False, name="Forecast band",
    ))
    fig.add_trace(go.Scatter(
        x=connector["date"], y=connector["forecast"],
        name="Naive forecast (no-change)", mode="lines",
        line=dict(color="#e67e22", width=2, dash="dash"),
        hovertemplate="<b>Forecast</b>: $%{y:.2f}<extra></extra>",
    ))
fig.update_layout(
    paper_bgcolor=BG, plot_bgcolor=BG,
    font=dict(color=TEXT, size=11),
    hovermode="x unified",
    showlegend=forecast_df is not None,
    legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1,
                font=dict(color=MUTED, size=10), bgcolor="rgba(0,0,0,0)"),
    margin=dict(l=55, r=20, t=15, b=40),
    xaxis=dict(
        **AXIS, title="",
        rangeselector=dict(
            buttons=[
                dict(count=1, label="1M", step="month", stepmode="backward"),
                dict(count=3, label="3M", step="month", stepmode="backward"),
                dict(count=6, label="6M", step="month", stepmode="backward"),
                dict(count=1, label="YTD", step="year", stepmode="todate"),
                dict(count=1, label="1Y", step="year", stepmode="backward"),
                dict(step="all", label="All"),
            ],
            bgcolor="#f6f8fa", activecolor=JPSI_BLUE,
            font=dict(color=TEXT, size=10), bordercolor=BORDER,
        ),
        rangeslider=dict(visible=False),
        type="date",
    ),
    yaxis=dict(**AXIS, title="$/cwt", tickprefix="$"),
    height=380,
)
add_watermark(fig, size=0.34, opacity=0.055)
st.plotly_chart(fig, use_container_width=True)
caption_bits = [
    "Line covers JSA's compiled workbook (published CME values) plus JSA's own USDA MARS "
    "reconstruction on both ends of that range — see the sidebar for methodology and accuracy notes."
]
if forecast_df is not None:
    caption_bits.append(
        f"Dashed orange segment is a {FORECAST_HORIZON_DAYS}-business-day naive (no-change) "
        f"projection with a {FORECAST_CI:.0%} band from historical day-ahead variability — backtested "
        "against a trend-following model and this simpler approach was actually more accurate (this "
        "series moves close to a random walk), but it's still not a trading signal or market forecast."
    )
st.caption(" ".join(caption_bits))


# ── Seasonal Pattern ──────────────────────────────────────────────────────────

st.markdown("<div style='height:18px'></div>", unsafe_allow_html=True)
st.markdown('<div class="sec-header">Seasonal Pattern by Year</div>', unsafe_allow_html=True)

seas = fci_df[["date", "fci_value"]].copy()
seas["year"] = seas["date"].dt.year
seas["doy"] = seas["date"].dt.dayofyear
current_year = seas["year"].max()

fig_seas = go.Figure()
for yr, grp in seas.groupby("year"):
    is_current = yr == current_year
    fig_seas.add_trace(go.Scatter(
        x=grp["doy"], y=grp["fci_value"],
        name=str(yr), mode="lines",
        line=dict(color=JPSI_BLUE if is_current else None, width=3 if is_current else 1.5),
        opacity=1.0 if is_current else 0.55,
        hovertemplate=f"<b>{yr}</b>: $%{{y:.2f}}<extra></extra>",
    ))
fig_seas.update_layout(
    paper_bgcolor=BG, plot_bgcolor=BG,
    font=dict(color=TEXT, size=11),
    hovermode="x unified",
    legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1,
                font=dict(color=MUTED, size=10), bgcolor="rgba(0,0,0,0)"),
    margin=dict(l=55, r=20, t=15, b=40),
    xaxis=dict(**AXIS, title="Day of Year"),
    yaxis=dict(**AXIS, title="$/cwt", tickprefix="$"),
    height=380,
)
add_watermark(fig_seas, size=0.3, opacity=0.06)
st.plotly_chart(fig_seas, use_container_width=True)
st.caption(
    f"Each line is one calendar year plotted by day-of-year ({current_year} bolded) — shows where "
    "this year sits against the same point in prior years. Not detrended: absolute levels differ "
    "year to year with broader market conditions, not just seasonality."
)


# ── Weekly Rundown ────────────────────────────────────────────────────────────

st.markdown("<div style='height:18px'></div>", unsafe_allow_html=True)
st.markdown('<div class="sec-header">Weekly Rundown</div>', unsafe_allow_html=True)

wk = fci_df.copy()
wk["week_end"] = wk["date"] - pd.to_timedelta(wk["date"].dt.weekday.map(lambda d: (d - 6) % 7), unit="D")
wk_summary = (
    wk.groupby("week_end")
    .agg(week_avg=("fci_value", "mean"), week_last=("fci_value", "last"),
         week_high=("fci_value", "max"), week_low=("fci_value", "min"))
    .reset_index()
    .sort_values("week_end")
)
wk_summary["prior_last"] = wk_summary["week_last"].shift(1)
wk_summary["week_chg"] = wk_summary["week_last"] - wk_summary["prior_last"]

display_wk = wk_summary.tail(10).sort_values("week_end", ascending=False).copy()
display_wk["Week Ending"] = display_wk["week_end"].dt.strftime("%m/%d/%Y")
display_wk = display_wk.rename(columns={
    "week_avg": "Week Avg", "week_last": "Week Last",
    "week_high": "Week High", "week_low": "Week Low", "week_chg": "Week Chg",
})[["Week Ending", "Week Avg", "Week Last", "Week High", "Week Low", "Week Chg"]]

with st.container(key="wm-weekly"):
    st.dataframe(
        display_wk.style.format({
            "Week Avg": "${:.2f}", "Week Last": "${:.2f}",
            "Week High": "${:.2f}", "Week Low": "${:.2f}", "Week Chg": "{:+.2f}",
        }, na_rep="—").map(
            lambda v: f"color: {POS}" if isinstance(v, (int, float)) and v > 0
            else (f"color: {NEG}" if isinstance(v, (int, float)) and v < 0 else ""),
            subset=["Week Chg"],
        ),
        use_container_width=True, hide_index=True, height=340,
    )


# ── Location Detail ───────────────────────────────────────────────────────────

st.markdown("<div style='height:18px'></div>", unsafe_allow_html=True)
st.markdown(
    f'<div class="sec-header">Sale Locations — {detail_date.strftime("%A, %B %d, %Y")}</div>',
    unsafe_allow_html=True,
)

day_rows = loc_filtered[loc_filtered["date"] == detail_date].sort_values("basis", ascending=False)
day_fci = value_on_or_before(fci_df, detail_date)

if day_rows.empty:
    st.info("No reporting sale locations on this date. Pick another date in the sidebar.")
else:
    left, right = st.columns([3, 2])
    with left:
        # Weighted totals row (matches how Compass's own report closes each
        # day's location table) -- weighted by head*weight, same formula as
        # the FCI itself, not a plain average of the Price column.
        total_head = day_rows["head"].sum()
        w = day_rows["head"] * day_rows["avg_weight"]
        total_weight = (w.sum() / total_head) if total_head else None
        total_price = ((w * day_rows["price"]).sum() / w.sum()) if w.sum() else None
        totals_row = pd.DataFrame([{
            "location": "TOTAL", "state": "",
            "head": total_head if total_head else pd.NA,
            "avg_weight": total_weight, "price": total_price,
            "basis": (total_price - day_fci) if (total_price is not None and day_fci is not None) else None,
        }])
        day_rows_with_total = pd.concat([day_rows, totals_row], ignore_index=True)

        disp = day_rows_with_total[["location", "state", "head", "avg_weight", "price", "basis"]].rename(columns={
            "location": "Location", "state": "State", "head": "Head", "avg_weight": "Weight",
            "price": "Price", "basis": "Basis vs FCI",
        })
        with st.container(key="wm-locations"):
            st.dataframe(
                disp.style.format({
                    "Head": "{:,.0f}", "Weight": "{:,.0f} lb", "Price": "${:.2f}", "Basis vs FCI": "{:+.2f}",
                }, na_rep="—").map(
                    lambda v: f"color: {POS}" if isinstance(v, (int, float)) and v > 0
                    else (f"color: {NEG}" if isinstance(v, (int, float)) and v < 0 else ""),
                    subset=["Basis vs FCI"],
                ).apply(
                    lambda row: ["font-weight:700;border-top:2px solid " + BORDER] * len(row)
                    if row["Location"] == "TOTAL" else [""] * len(row),
                    axis=1,
                ),
                use_container_width=True, hide_index=True, height=380,
            )
    with right:
        fig_b = go.Figure()
        bar_colors = [POS if v >= 0 else NEG for v in day_rows["basis"]]
        fig_b.add_trace(go.Bar(
            x=day_rows["basis"], y=day_rows["location"],
            orientation="h", marker_color=bar_colors,
            hovertemplate="<b>%{y}</b>: %{x:+.2f}<extra></extra>",
        ))
        fig_b.update_layout(
            paper_bgcolor=BG, plot_bgcolor=BG,
            font=dict(color=TEXT, size=10),
            margin=dict(l=10, r=10, t=10, b=30),
            xaxis=dict(**AXIS, title="Basis vs FCI ($/cwt)"),
            yaxis=dict(**AXIS, autorange="reversed"),
            height=380, showlegend=False,
        )
        add_watermark(fig_b, size=0.4, opacity=0.06)
        st.plotly_chart(fig_b, use_container_width=True)
    if day_fci is not None:
        st.caption(f"Index value used for basis: **${day_fci:.2f}**")


# ── Basis Leaderboard ─────────────────────────────────────────────────────────

st.markdown("<div style='height:18px'></div>", unsafe_allow_html=True)
st.markdown('<div class="sec-header">Average Basis by Location (Trailing 90 Days)</div>', unsafe_allow_html=True)

window_start = last_date - timedelta(days=90)
recent = loc_filtered[loc_filtered["date"] >= window_start]
leaderboard = (
    recent.groupby("location")
    .agg(avg_basis=("basis", "mean"), sales=("basis", "size"))
    .query("sales >= 3")
    .sort_values("avg_basis", ascending=False)
    .reset_index()
)

if leaderboard.empty:
    st.info("Not enough recent sales to build a basis leaderboard.")
else:
    top_bottom = pd.concat([leaderboard.head(10), leaderboard.tail(10)]).drop_duplicates(subset="location")
    top_bottom = top_bottom.sort_values("avg_basis", ascending=True)
    fig_lb = go.Figure()
    lb_colors = [POS if v >= 0 else NEG for v in top_bottom["avg_basis"]]
    fig_lb.add_trace(go.Bar(
        x=top_bottom["avg_basis"], y=top_bottom["location"],
        orientation="h", marker_color=lb_colors,
        hovertemplate="<b>%{y}</b>: %{x:+.2f} avg basis<extra></extra>",
    ))
    fig_lb.update_layout(
        paper_bgcolor=BG, plot_bgcolor=BG,
        font=dict(color=TEXT, size=11),
        margin=dict(l=10, r=10, t=10, b=30),
        xaxis=dict(**AXIS, title="Avg Basis vs FCI ($/cwt)"),
        yaxis=dict(**AXIS),
        height=440, showlegend=False,
    )
    add_watermark(fig_lb, size=0.3, opacity=0.06)
    st.plotly_chart(fig_lb, use_container_width=True)
    st.caption("Locations with at least 3 reported sales in the trailing 90 days, strongest and weakest basis shown.")


# ── Data Table ────────────────────────────────────────────────────────────────

with st.expander("📋  Raw Data Table"):
    tab_fci, tab_loc = st.tabs(["Index Values", "Location Sales"])
    with tab_fci:
        d = fci_df.copy()
        d["date"] = d["date"].dt.strftime("%Y-%m-%d")
        with st.container(key="wm-raw-fci"):
            st.dataframe(
                d.rename(columns={"date": "Date", "fci_value": "FCI"}).sort_values("Date", ascending=False)
                .style.format({"FCI": "${:.2f}"}),
                use_container_width=True, hide_index=True, height=320,
            )
    with tab_loc:
        d = loc_filtered[["date", "location", "state", "head", "avg_weight", "price", "fci_value", "basis"]].copy()
        d["date"] = d["date"].dt.strftime("%Y-%m-%d")
        with st.container(key="wm-raw-loc"):
            st.dataframe(
                d.rename(columns={
                    "date": "Date", "location": "Location", "state": "State", "head": "Head",
                    "avg_weight": "Weight", "price": "Price", "fci_value": "FCI", "basis": "Basis",
                }).sort_values("Date", ascending=False)
                .style.format({
                    "Head": "{:,.0f}", "Weight": "{:,.0f} lb", "Price": "${:.2f}", "FCI": "${:.2f}", "Basis": "{:+.2f}",
                }, na_rep="—"),
                use_container_width=True, hide_index=True, height=320,
            )


# ── Footer ────────────────────────────────────────────────────────────────────

_year = datetime.now().year
st.markdown(f"<hr style='border-color:{BORDER};margin-top:32px;margin-bottom:16px'>", unsafe_allow_html=True)
st.markdown(
    f'<div style="color:{MUTED};font-size:0.68rem;line-height:1.6;text-align:center;padding:0 24px 24px;">'
    f'Historical index and basis figures are derived from JSA-compiled 12-state feeder steer sale data '
    f'(coverage: {first_date.strftime("%b %d, %Y")}–{last_date.strftime("%b %d, %Y")}) and are provided for informational purposes only. '
    f'Trading commodity futures, options on futures, cash commodities, and over-the-counter derivative products involves substantial risk of loss and may not be suitable for all investors. '
    f'This communication does not constitute investment advice, a recommendation, or an offer or solicitation to buy or sell any futures, options, cash commodities, or derivative products. '
    f'John Stewart &amp; Associates, Inc. does not accept orders to buy or sell any financial instruments via email. '
    f'The information contained herein has been obtained from sources believed to be reliable; however, its accuracy and completeness are not guaranteed. '
    f'&copy; John Stewart &amp; Associates, Inc. {_year}'
    f'</div>',
    unsafe_allow_html=True,
)
